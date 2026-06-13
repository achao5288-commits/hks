"""
Excel generation executor using openpyxl.
Creates .xlsx files with data tables and optional charts.
"""
import os
import time
import tempfile
from .base import BaseExecutor, WorkflowExecutionError


class ExcelChartExecutor(BaseExecutor):

    def get_config_schema(self) -> dict:
        return {
            "type": "object",
            "title": "Excel生成配置",
            "properties": {
                "output_path": {
                    "type": "string", "title": "输出路径",
                    "description": "Excel文件保存路径，留空使用临时目录", "default": "",
                },
                "sheet_name": {"type": "string", "title": "工作表名称", "default": "Sheet1"},
                "create_chart": {"type": "boolean", "title": "生成图表", "default": False},
                "chart_type": {
                    "type": "string", "title": "图表类型",
                    "enum": ["bar", "line", "pie"], "default": "bar",
                },
                "chart_title": {"type": "string", "title": "图表标题", "default": "数据图表"},
                "chart_category_column": {
                    "type": "string", "title": "图表分类列",
                    "description": "用作X轴/分类的列名", "default": "",
                },
                "chart_value_column": {
                    "type": "string", "title": "图表数值列",
                    "description": "用作Y轴/数值的列名", "default": "",
                },
                "column_widths": {
                    "type": "object", "title": "列宽设置",
                    "description": "列名到宽度的映射，如 {'标题': 30} 或 {'A': 20}",
                },
                "auto_fit_columns": {"type": "boolean", "title": "自适应列宽", "default": True},
            },
            "required": [],
        }

    def execute(self, config: dict, context: dict) -> dict:
        import traceback as _tb
        try:
            import openpyxl
            from openpyxl.chart import BarChart, LineChart, PieChart, Reference
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise WorkflowExecutionError("DEP_MISSING", "openpyxl is not installed", "excel_chart")

        rows = self._collect_rows(context)
        if not rows:
            return {
                "status": "success",
                "data": {"file_path": "", "file_size": 0, "sheet_count": 0, "message": "No data to write"},
            }

        headers = list(rows[0].keys()) if rows else []
        wb = openpyxl.Workbook()
        ws = wb.active

        try:
            self._build_sheet(ws, rows, headers, config, get_column_letter, openpyxl)
        except Exception as e:
            import sys, traceback as _tb2
            _tb2.print_exc(file=sys.stderr)
            sys.stderr.flush()
            raise WorkflowExecutionError("EXCEL_ERROR", f"{e}\n{_tb2.format_exc()}", "excel_chart")

        try:
            output_path = config.get("output_path", "")
            if not output_path:
                output_path = os.path.join(tempfile.gettempdir(), f"workflow_report_{int(time.time())}.xlsx")

            # Sanitize: remove any column_dimensions with non-letter keys before save
            bad_keys = [k for k in list(ws.column_dimensions)
                        if not (str(k).isascii() and str(k).isalpha() and len(str(k)) <= 3)]
            if bad_keys:
                import sys as _s
                print(f'[excel_chart] Purging bad column keys: {bad_keys}', file=_s.stderr)
                _s.stderr.flush()
                for key in bad_keys:
                    try:
                        del ws.column_dimensions[key]
                    except Exception:
                        pass

            os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
            wb.save(output_path)
            file_size = os.path.getsize(output_path) if os.path.exists(output_path) else 0

            return {
                "status": "success",
                "data": {
                    "file_path": output_path, "file_size": file_size,
                    "sheet_count": len(wb.sheetnames), "row_count": len(rows), "column_count": len(headers),
                },
            }
        except Exception as e2:
            import sys as _sys2, traceback as _tb3
            _tb3.print_exc(file=_sys2.stderr)
            _sys2.stderr.flush()
            raise WorkflowExecutionError("EXCEL_SAVE_ERROR", f"{e2}\n{_tb3.format_exc()}", "excel_chart")

    def _build_sheet(self, ws, rows, headers, config, get_column_letter, openpyxl):
        """Build the sheet: headers, data, column widths, optional chart."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        ws.title = config.get("sheet_name", "Sheet1")

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, row in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(header, "")).border = thin_border

        # Auto-fit columns (always safe — uses get_column_letter with integer)
        if config.get("auto_fit_columns", True):
            for col_idx, header in enumerate(headers, 1):
                max_width = len(str(header)) + 2
                for row in rows:
                    max_width = max(max_width, len(str(row.get(header, ""))))
                try:
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width + 2, 50)
                except Exception:
                    pass

        # Custom column widths — support BOTH column names AND column letters
        column_widths = config.get("column_widths", {})
        if isinstance(column_widths, dict):
            for col_key, width in column_widths.items():
                try:
                    letter = col_key
                    if not (col_key.isascii() and col_key.isalpha() and len(col_key) <= 3):
                        if col_key in headers:
                            letter = get_column_letter(headers.index(col_key) + 1)
                    ws.column_dimensions[letter].width = int(width)
                except Exception:
                    pass

        # Optional chart
        if config.get("create_chart") and len(rows) > 0:
            self._add_chart(ws, rows, headers, config, get_column_letter, openpyxl)

    def _add_chart(self, ws, rows, headers, config, get_column_letter, openpyxl):
        """Add a chart to the worksheet."""
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference

        chart_type = config.get("chart_type", "bar")
        chart_title = config.get("chart_title", "数据图表")
        cat_col = config.get("chart_category_column", headers[0] if headers else "")
        val_col = config.get("chart_value_column", "")

        # Resolve column names to indices with safe fallbacks
        def _col_idx(name, default):
            if name and name in headers:
                return headers.index(name) + 1
            return default

        cat_idx = _col_idx(cat_col, 1)
        val_idx = _col_idx(val_col, 2 if len(headers) > 1 else 1)
        last_row = len(rows) + 1

        chart_map = {"bar": BarChart, "line": LineChart, "pie": PieChart}
        chart_class = chart_map.get(chart_type, BarChart)
        chart = chart_class()
        chart.title = chart_title
        chart.style = 10

        data_ref = Reference(ws, min_col=val_idx, min_row=1, max_row=last_row, max_col=val_idx)
        cats_ref = Reference(ws, min_col=cat_idx, min_row=2, max_row=last_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        ws.add_chart(chart, f"A{last_row + 3}")

    def _collect_rows(self, context: dict) -> list:
        all_rows = []
        for node_id, result in context.items():
            data = result.get("data", result)
            if isinstance(data, dict):
                for key in ("rows", "items", "selector_results"):
                    if key in data:
                        if key == "selector_results":
                            sr = data[key]
                            list_keys = [k for k, v in sr.items() if isinstance(v, list)]
                            if list_keys:
                                max_len = max((len(sr[k]) for k in list_keys), default=0)
                                for i in range(max_len):
                                    row = {}
                                    for k in list_keys:
                                        if i < len(sr[k]):
                                            row[k] = sr[k][i]
                                    if row:
                                        all_rows.append(row)
                            else:
                                row = {k: v for k, v in sr.items() if not k.endswith("_error")}
                                if row:
                                    all_rows.append(row)
                        else:
                            items = data[key]
                            if isinstance(items, list):
                                all_rows.extend(items)
                        break
                else:
                    all_rows.append(data)
            elif isinstance(data, list):
                all_rows.extend(data)
        return all_rows
